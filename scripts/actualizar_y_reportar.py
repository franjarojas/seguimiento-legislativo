"""
Actualiza el Excel de seguimiento legislativo y envía reporte de cambios por correo.

Flujo:
  1. Carga estado_anterior.json (snapshot de la semana pasada)
  2. Scraping: actualiza Col3, Col4, Col6 de cada boletín via tramitacion.senado.cl
  3. Compara con estado anterior → detecta cambios
  4. Guarda estado_actual.json (pasa a ser el nuevo estado_anterior.json)
  5. Si hay cambios (o es la primera ejecución): envía correo HTML con tabla de cambios + Excel adjunto

Variables de entorno requeridas (GitHub Secrets):
  GMAIL_TOKEN_JSON   — contenido del token OAuth de Gmail (ver README)
  DESTINATARIOS      — emails separados por coma, ej: "a@x.cl,b@x.cl"

Uso local:
  pip install requests openpyxl beautifulsoup4 google-auth google-auth-oauthlib google-api-python-client
  python scripts/actualizar_y_reportar.py
  python scripts/actualizar_y_reportar.py --solo-reporte   # no hace scraping, solo envía correo
"""

import os
import re
import sys
import json
import time
import base64
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Rutas ──────────────────────────────────────────────────────────────────
ROOT            = Path(__file__).parent.parent
EXCEL_PATH      = ROOT / "seguimiento_legislativo.xlsx"
ESTADO_PATH     = ROOT / "estado_anterior.json"

# ── Configuración ──────────────────────────────────────────────────────────
SEMANAS_ALERTA  = 4   # marcar en el correo si un proyecto lleva N semanas sin cambios
BASE_SENADO     = "https://tramitacion.senado.cl/appsenado/index.php"
HEADERS         = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://tramitacion.senado.cl/appsenado/templates/tramitacion/index.php",
    "X-Requested-With": "XMLHttpRequest",
}

# Índices de columna (1-based)
COL_BOLETIN   = 1
COL_NOMBRE    = 2
COL_FECHA_ULT = 3
COL_ULT_MOV   = 4
COL_URGENCIA  = 5
COL_UBICACION = 6
COL_COMISION  = 7
COL_FECHA_PRS = 8
COL_ESTADO    = 9
COL_TIPO      = 10
COL_CAMARA    = 11
COL_AUTORES   = 12


# ══════════════════════════════════════════════════════════════════════════════
# SCRAPING
# ══════════════════════════════════════════════════════════════════════════════

def ts():
    return int(time.time() * 1000)


def _parse_fecha(fecha_str) -> datetime.datetime:
    try:
        return datetime.datetime.strptime(str(fecha_str or ""), "%d/%m/%Y")
    except Exception:
        return datetime.datetime.min


def normalizar_boletin(val: str) -> str:
    val = val.strip().replace(".", "")
    if "-" not in val and len(val) > 2:
        val = val[:-2] + "-" + val[-2:]
    return val


def obtener_proyid(boletin: str) -> str | None:
    params = {"mo": "tramitacion", "ac": "datos_proy", "nboletin": boletin, "etc": ts()}
    try:
        r = requests.get(BASE_SENADO, params=params, headers=HEADERS, timeout=15)
        m = re.search(r'proyid\s*=\s*(\d+)', r.text)
        return m.group(1) if m else None
    except Exception:
        return None


def obtener_ultimo_tramite(proyid: str, boletin: str) -> dict:
    params = {
        "mo": "tramitacion", "ac": "tramites",
        "proyid": proyid, "camara": "S",
        "boletin": boletin, "etc": ts()
    }
    try:
        r = requests.get(BASE_SENADO, params=params, headers=HEADERS, timeout=15)
    except Exception as e:
        return {"error": str(e)}

    soup = BeautifulSoup(r.text, "html.parser")
    tabla = soup.find("table", id="grid_tram") or soup.find("table")
    if not tabla:
        return {"error": "No se encontró tabla"}

    filas_datos = []
    for fila in tabla.find_all("tr"):
        celdas = fila.find_all("td")
        if len(celdas) >= 3:
            textos = [c.get_text(strip=True) for c in celdas]
            if len(textos) > 1 and re.match(r'\d{2}/\d{2}/\d{4}', textos[1]):
                filas_datos.append(textos)

    if not filas_datos:
        return {"error": "No se encontraron filas con fechas"}

    ultimo = filas_datos[-1]
    fecha  = ultimo[1] if len(ultimo) > 1 else "—"
    desc   = ultimo[2] if len(ultimo) > 2 else "—"
    etapa  = ultimo[3] if len(ultimo) > 3 else ""
    camara = ultimo[4] if len(ultimo) > 4 else ""

    estado = " — ".join(p for p in [etapa, camara] if p) or etapa
    descripcion = " — ".join(p for p in [etapa, desc] if p) if etapa else desc

    return {"fecha": fecha, "descripcion": descripcion, "estado": estado}


def obtener_datos_proy(boletin: str) -> dict:
    """Obtiene metadatos del proyecto: título, fecha ingreso, cámara, iniciativa, urgencia."""
    params = {"mo": "tramitacion", "ac": "datos_proy", "nboletin": boletin, "etc": ts()}
    resultado = {}
    try:
        r = requests.get(BASE_SENADO, params=params, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        tabla = soup.find("table", class_="table-datos")
        if tabla:
            rows = tabla.find_all("tr")
            for i, tr in enumerate(rows):
                tds = tr.find_all("td")
                vals = [td.get_text(strip=True) for td in tds]
                if not vals:
                    continue
                # La tabla tiene 4 columnas por fila (2 pares etiqueta-valor).
                # Procesamos ambos pares: (vals[0], vals[1]) y (vals[2], vals[3]).
                pares = [(vals[0], vals[1] if len(vals) > 1 else "")]
                if len(vals) >= 4:
                    pares.append((vals[2], vals[3]))
                for label_raw, valor in pares:
                    label = label_raw.rstrip(":").strip()
                    if label == "Título":
                        resultado["nombre"] = valor
                    elif label == "Fecha de Ingreso":
                        resultado["fecha_prs"] = valor
                    elif label in ("Cámara de Origen", "Camara de Origen"):
                        resultado["camara"] = valor
                    elif label == "Iniciativa":
                        resultado["tipo"] = valor
                    elif label == "Urgencia Actual":
                        resultado["urgencia"] = valor
                # Etapa / subetapa sólo en el par izquierdo
                label0 = vals[0].rstrip(":").strip()
                if label0 in ("Etapa:", "Etapa"):
                    if i + 1 < len(rows):
                        next_tds = rows[i + 1].find_all("td")
                        if len(next_tds) > 1:
                            resultado["subetapa"] = next_tds[1].get_text(strip=True)
                        elif next_tds:
                            resultado["subetapa"] = next_tds[0].get_text(strip=True)
    except Exception:
        pass
    return resultado


def obtener_autores(proyid: str) -> str:
    """Devuelve los autores del proyecto como string separado por '; '."""
    params = {"mo": "tramitacion", "ac": "autores", "proyid": proyid, "etc": ts()}
    try:
        r = requests.get(BASE_SENADO, params=params, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        tabla = soup.find("table", id="grid_aut") or soup.find("table")
        if not tabla:
            return "—"
        autores = []
        for fila in tabla.find_all("tr"):
            celdas = fila.find_all("td")
            if len(celdas) >= 2:
                nombre = celdas[1].get_text(strip=True)
                if nombre and nombre != "Autor":
                    autores.append(nombre)
        return "; ".join(autores) if autores else "—"
    except Exception:
        return "—"


def derivar_comision(subetapa: str) -> str:
    """Extrae la comisión actual o devuelve 'En Sala' si el proyecto está en sala."""
    s = subetapa.strip()
    m = re.search(r'informe de comisi[oó]n de (.+)', s, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    if re.search(r'discusi[oó]n', s, re.IGNORECASE):
        return "En Sala"
    return s if s else "—"


def derivar_ubicacion(estado: str) -> str:
    """Extrae la ubicación actual (Senado / Cámara) desde el estado de tramitación."""
    e = estado.lower()
    if "senado" in e:
        return "Senado"
    if "diputados" in e or "diputado" in e:
        return "Cámara"
    if "mixta" in e:
        return "Comisión Mixta"
    return "—"


def obtener_urgencia(boletin: str) -> str:
    datos = obtener_datos_proy(boletin)
    return datos.get("urgencia", "—")


def scraping_boletin(boletin: str, completar_estaticos: bool = False) -> dict:
    proyid = obtener_proyid(boletin)
    if not proyid:
        return {"error": "No se pudo obtener proyid"}
    time.sleep(0.3)
    resultado = obtener_ultimo_tramite(proyid, boletin)
    if "error" in resultado:
        return resultado

    # Urgencia y metadatos estáticos desde datos_proy (una sola llamada)
    datos = obtener_datos_proy(boletin)
    resultado["urgencia"]  = datos.get("urgencia", "—")
    resultado["ubicacion"] = derivar_ubicacion(resultado.get("estado", ""))
    resultado["comision"]  = derivar_comision(datos.get("subetapa", ""))

    if completar_estaticos:
        time.sleep(0.3)
        resultado["nombre"]    = datos.get("nombre", "")
        resultado["fecha_prs"] = datos.get("fecha_prs", "")
        resultado["camara"]    = datos.get("camara", "")
        resultado["tipo"]      = datos.get("tipo", "")
        resultado["autores"]   = obtener_autores(proyid)

    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════

# Columnas centradas vs. con wrap izquierdo
COLS_CENTRADAS = {COL_BOLETIN, COL_FECHA_ULT, COL_URGENCIA, COL_UBICACION,
                  COL_FECHA_PRS, COL_TIPO, COL_CAMARA}
FILL_ALTERNO   = PatternFill("solid", fgColor="EBF3FB")
BORDE_CELDA    = Border(
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)


def estandarizar_formato(ws) -> None:
    """Aplica diseño uniforme a todas las celdas de datos (fila 2 en adelante)."""
    font_d = Font(name="Arial", size=9)
    for row in ws.iter_rows(min_row=2):
        # Filas pares (índice 0-based impar) → fondo alterno azul claro
        usar_fill = (row[0].row % 2 == 1)  # fila 3, 5, 7… → True
        for cell in row:
            cell.font      = font_d
            cell.border    = BORDE_CELDA
            cell.fill      = FILL_ALTERNO if usar_fill else PatternFill()
            horiz = "center" if cell.column in COLS_CENTRADAS else "left"
            cell.alignment = Alignment(horizontal=horiz, vertical="top", wrap_text=True)


def leer_estado_excel(ruta: Path) -> dict:
    """Devuelve dict {boletin: {nombre, fecha_ult, ult_mov, estado}} para comparación."""
    wb = load_workbook(ruta)
    ws = wb.active
    estado = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        boletin = str(row[COL_BOLETIN - 1] or "").strip()
        if not boletin or boletin == "—":
            continue
        estado[boletin] = {
            "nombre":    str(row[COL_NOMBRE - 1]    or ""),
            "fecha_ult": str(row[COL_FECHA_ULT - 1] or ""),
            "ult_mov":   str(row[COL_ULT_MOV - 1]   or ""),
            "urgencia":  str(row[COL_URGENCIA - 1]  or ""),
            "ubicacion": str(row[COL_UBICACION - 1] or ""),
            "comision":  str(row[COL_COMISION - 1]  or ""),
            "estado":    str(row[COL_ESTADO - 1]     or ""),
        }
    return estado


def actualizar_excel(ruta: Path) -> dict:
    """Corre scraping y actualiza el Excel. Devuelve {boletin: resultado_scraping}."""
    wb = load_workbook(ruta)
    ws = wb.active
    font_d = Font(name="Arial", size=9)
    a_wrap = Alignment(vertical="top", wrap_text=True)
    a_ctr  = Alignment(horizontal="center", vertical="top", wrap_text=True)

    resultados = {}
    for row in ws.iter_rows(min_row=2):
        val_raw = str(row[COL_BOLETIN - 1].value or "").strip()
        if not val_raw or val_raw == "—":
            continue
        boletin = normalizar_boletin(val_raw)
        # Completar datos estáticos si el nombre está vacío o es genérico
        nombre_actual = str(row[COL_NOMBRE - 1].value or "").strip()
        completar = (not nombre_actual or
                     nombre_actual in ("—", "Sin información disponible") or
                     "no indexado" in nombre_actual.lower())
        print(f"  {boletin}{'[completar]' if completar else ''}...", end=" ", flush=True)
        r = scraping_boletin(boletin, completar_estaticos=completar)
        resultados[val_raw] = r

        if "error" in r:
            print(f"ERROR: {r['error']}")
            continue

        row[COL_FECHA_ULT - 1].value     = r["fecha"]
        row[COL_FECHA_ULT - 1].font      = font_d
        row[COL_FECHA_ULT - 1].alignment = a_ctr
        row[COL_ULT_MOV - 1].value       = r["descripcion"]
        row[COL_ULT_MOV - 1].font        = font_d
        row[COL_ULT_MOV - 1].alignment   = a_wrap
        row[COL_URGENCIA - 1].value      = r.get("urgencia", "—")
        row[COL_URGENCIA - 1].font       = font_d
        row[COL_URGENCIA - 1].alignment  = a_ctr
        row[COL_UBICACION - 1].value     = r.get("ubicacion", "—")
        row[COL_UBICACION - 1].font      = font_d
        row[COL_UBICACION - 1].alignment = a_ctr
        row[COL_COMISION - 1].value      = r.get("comision", "—")
        row[COL_COMISION - 1].font       = font_d
        row[COL_COMISION - 1].alignment  = a_wrap
        row[COL_ESTADO - 1].value         = r["estado"]
        row[COL_ESTADO - 1].font          = font_d
        row[COL_ESTADO - 1].alignment     = a_wrap
        # Rellenar datos estáticos si fueron obtenidos
        if r.get("nombre"):
            row[COL_NOMBRE - 1].value     = r["nombre"]
            row[COL_NOMBRE - 1].font      = font_d
            row[COL_NOMBRE - 1].alignment = a_wrap
        if r.get("fecha_prs"):
            row[COL_FECHA_PRS - 1].value     = r["fecha_prs"]
            row[COL_FECHA_PRS - 1].font      = font_d
            row[COL_FECHA_PRS - 1].alignment = a_wrap
        if r.get("tipo"):
            row[COL_TIPO - 1].value     = r["tipo"]
            row[COL_TIPO - 1].font      = font_d
            row[COL_TIPO - 1].alignment = a_ctr
        if r.get("camara"):
            row[COL_CAMARA - 1].value     = r["camara"]
            row[COL_CAMARA - 1].font      = font_d
            row[COL_CAMARA - 1].alignment = a_ctr
        if r.get("autores"):
            row[COL_AUTORES - 1].value     = r["autores"]
            row[COL_AUTORES - 1].font      = font_d
            row[COL_AUTORES - 1].alignment = a_wrap
        print(f"OK → {r['fecha']} | urgencia: {r.get('urgencia', '—')}")

    # Ordenar filas por fecha de último movimiento (más reciente primero)
    filas = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if any(cell.value for cell in row):
            filas.append([cell.value for cell in row])
    filas.sort(key=lambda f: _parse_fecha(f[COL_FECHA_ULT - 1]), reverse=True)
    for i, fila in enumerate(filas, start=2):
        for j, valor in enumerate(fila, start=1):
            ws.cell(row=i, column=j, value=valor)

    estandarizar_formato(ws)
    wb.save(ruta)
    return resultados


# ══════════════════════════════════════════════════════════════════════════════
# COMPARACIÓN DE CAMBIOS
# ══════════════════════════════════════════════════════════════════════════════

def detectar_cambios(estado_anterior: dict, estado_nuevo: dict) -> list[dict]:
    """
    Retorna lista de boletines con cambios.
    Cada item: {boletin, nombre, campo_cambiado, valor_anterior, valor_nuevo, fecha_ult, semanas_sin_cambio}
    """
    cambios = []
    hoy = datetime.date.today()

    for boletin, nuevo in estado_nuevo.items():
        anterior = estado_anterior.get(boletin, {})

        # Calcular semanas sin cambio
        semanas_sin_cambio = None
        try:
            d = datetime.datetime.strptime(nuevo["fecha_ult"], "%d/%m/%Y").date()
            semanas_sin_cambio = (hoy - d).days // 7
        except Exception:
            pass

        cambio = {
            "boletin": boletin,
            "nombre": nuevo["nombre"][:70] + "…" if len(nuevo["nombre"]) > 70 else nuevo["nombre"],
            "fecha_ult": nuevo["fecha_ult"],
            "ult_mov": nuevo["ult_mov"],
            "estado_nuevo": nuevo["estado"],
            "estado_anterior": anterior.get("estado", "—"),
            "semanas_sin_cambio": semanas_sin_cambio,
            "es_nuevo": boletin not in estado_anterior,
        }

        # ¿Hubo cambio en los campos de tramitación?
        if anterior:
            hubo_cambio = (
                anterior.get("fecha_ult") != nuevo["fecha_ult"] or
                anterior.get("ult_mov")   != nuevo["ult_mov"]   or
                anterior.get("estado")    != nuevo["estado"]
            )
            cambio["hubo_cambio"] = hubo_cambio
        else:
            cambio["hubo_cambio"] = True  # nuevo boletín

        cambios.append(cambio)

    return cambios


# ══════════════════════════════════════════════════════════════════════════════
# CORREO HTML
# ══════════════════════════════════════════════════════════════════════════════

SEMANAS_RECIENTE = 4  # destacar proyectos con movimiento en las últimas N semanas


def construir_html(cambios: list[dict], fecha_ejecucion: str) -> str:
    solo_cambios   = sorted([c for c in cambios if c["hubo_cambio"]],
                            key=lambda c: _parse_fecha(c["fecha_ult"]), reverse=True)
    sin_cambio     = sorted([c for c in cambios if not c["hubo_cambio"]],
                            key=lambda c: _parse_fecha(c["fecha_ult"]), reverse=True)

    # Proyectos con movimiento reciente (últimas 2 semanas), excluyendo los que
    # ya aparecen en "cambios esta semana"
    con_movimiento_reciente = [
        c for c in sin_cambio
        if c["semanas_sin_cambio"] is not None
        and c["semanas_sin_cambio"] <= SEMANAS_RECIENTE
    ]

    con_alerta = [
        c for c in sin_cambio
        if c["semanas_sin_cambio"] is not None
        and c["semanas_sin_cambio"] >= SEMANAS_ALERTA
    ]

    def fila_cambio(c):
        estado_ant = c["estado_anterior"] if c["estado_anterior"] != "—" else "<em>nuevo</em>"
        flecha = "🆕" if c["es_nuevo"] else "🔄"
        return f"""
        <tr>
          <td style="padding:8px;border-bottom:1px solid #e2e8f0;font-weight:600;white-space:nowrap">{flecha} {c['boletin']}</td>
          <td style="padding:8px;border-bottom:1px solid #e2e8f0;font-size:13px">{c['nombre']}</td>
          <td style="padding:8px;border-bottom:1px solid #e2e8f0;white-space:nowrap;color:#64748b">{c['fecha_ult']}</td>
          <td style="padding:8px;border-bottom:1px solid #e2e8f0;font-size:13px;color:#dc2626">{estado_ant}</td>
          <td style="padding:8px;border-bottom:1px solid #e2e8f0;font-size:13px;color:#16a34a">{c['estado_nuevo']}</td>
        </tr>"""

    def fila_reciente(c):
        return f"""
        <tr>
          <td style="padding:8px;border-bottom:1px solid #dcfce7;font-weight:600;white-space:nowrap">🟢 {c['boletin']}</td>
          <td style="padding:8px;border-bottom:1px solid #dcfce7;font-size:13px">{c['nombre']}</td>
          <td style="padding:8px;border-bottom:1px solid #dcfce7;white-space:nowrap">{c['fecha_ult']}</td>
          <td style="padding:8px;border-bottom:1px solid #dcfce7;font-size:13px">{c['ult_mov']}</td>
          <td style="padding:8px;border-bottom:1px solid #dcfce7;color:#15803d">
            Hace {c['semanas_sin_cambio']} semana{'s' if c['semanas_sin_cambio'] != 1 else ''}
          </td>
        </tr>"""

    def fila_alerta(c):
        return f"""
        <tr>
          <td style="padding:8px;border-bottom:1px solid #fef3c7;font-weight:600;white-space:nowrap">⚠️ {c['boletin']}</td>
          <td style="padding:8px;border-bottom:1px solid #fef3c7;font-size:13px">{c['nombre']}</td>
          <td style="padding:8px;border-bottom:1px solid #fef3c7;white-space:nowrap">{c['fecha_ult']}</td>
          <td style="padding:8px;border-bottom:1px solid #fef3c7;color:#92400e">{c['semanas_sin_cambio']} semanas sin movimiento</td>
        </tr>"""

    # ── Sección cambios esta semana ──
    tabla_cambios = ""
    if solo_cambios:
        filas = "".join(fila_cambio(c) for c in solo_cambios)
        tabla_cambios = f"""
        <h2 style="color:#1e3a5f;font-size:16px;margin-top:28px">
          🔄 Boletines con cambios esta semana ({len(solo_cambios)})
        </h2>
        <table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px">
          <thead>
            <tr style="background:#1e3a5f;color:white">
              <th style="padding:10px;text-align:left">Boletín</th>
              <th style="padding:10px;text-align:left">Nombre</th>
              <th style="padding:10px;text-align:left">Fecha</th>
              <th style="padding:10px;text-align:left">Estado anterior</th>
              <th style="padding:10px;text-align:left">Estado nuevo</th>
            </tr>
          </thead>
          <tbody>{filas}</tbody>
        </table>"""
    else:
        tabla_cambios = """
        <p style="color:#64748b;font-style:italic;margin-top:20px">
          ✅ No hubo cambios de tramitación esta semana.
        </p>"""

    # ── Sección movimiento reciente (últimas 2 semanas) ──
    tabla_reciente = ""
    if con_movimiento_reciente:
        filas_r = "".join(fila_reciente(c) for c in con_movimiento_reciente)
        tabla_reciente = f"""
        <h2 style="color:#15803d;font-size:16px;margin-top:28px">
          🟢 Con movimiento reciente — último mes ({len(con_movimiento_reciente)})
        </h2>
        <table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;background:#f0fdf4">
          <thead>
            <tr style="background:#16a34a;color:white">
              <th style="padding:10px;text-align:left">Boletín</th>
              <th style="padding:10px;text-align:left">Nombre</th>
              <th style="padding:10px;text-align:left">Fecha</th>
              <th style="padding:10px;text-align:left">Último movimiento</th>
              <th style="padding:10px;text-align:left">Antigüedad</th>
            </tr>
          </thead>
          <tbody>{filas_r}</tbody>
        </table>"""

    # ── Sección alertas de inactividad ──
    tabla_alertas = ""
    if con_alerta:
        filas_a = "".join(fila_alerta(c) for c in con_alerta)
        tabla_alertas = f"""
        <h2 style="color:#92400e;font-size:16px;margin-top:28px">
          ⚠️ Proyectos sin movimiento hace {SEMANAS_ALERTA}+ semanas ({len(con_alerta)})
        </h2>
        <table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;background:#fffbeb">
          <thead>
            <tr style="background:#fbbf24;color:#1c1917">
              <th style="padding:10px;text-align:left">Boletín</th>
              <th style="padding:10px;text-align:left">Nombre</th>
              <th style="padding:10px;text-align:left">Último movimiento</th>
              <th style="padding:10px;text-align:left">Inactividad</th>
            </tr>
          </thead>
          <tbody>{filas_a}</tbody>
        </table>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;max-width:860px;margin:0 auto;padding:20px;color:#1e293b">
  <div style="background:#1e3a5f;padding:20px 28px;border-radius:8px 8px 0 0">
    <h1 style="color:white;margin:0;font-size:20px">
      📋 Reporte Seguimiento Legislativo
    </h1>
    <p style="color:#93c5fd;margin:6px 0 0">
      Actualización semanal — {fecha_ejecucion}
    </p>
  </div>
  <div style="background:#f8fafc;padding:20px 28px;border:1px solid #e2e8f0;border-top:none;border-radius:0 0 8px 8px">
    <p style="margin:0;font-size:14px;color:#475569">
      Se monitorearon <strong>{len(cambios)} boletines</strong>.
      Adjunto encontrarás el Excel actualizado.
    </p>
    {tabla_cambios}
    {tabla_reciente}
    {tabla_alertas}
    <hr style="border:none;border-top:1px solid #e2e8f0;margin-top:32px">
    <p style="font-size:12px;color:#94a3b8;margin:12px 0 0">
      Generado automáticamente por el sistema de seguimiento legislativo del Proyecto REDAR.
      Fuente: tramitacion.senado.cl / camara.cl
    </p>
  </div>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
# ENVÍO POR GMAIL API
# ══════════════════════════════════════════════════════════════════════════════

def enviar_correo(html_body: str, excel_path: Path, destinatarios: list[str], fecha: str):
    import google.oauth2.credentials
    from googleapiclient.discovery import build

    token_json = os.environ.get("GMAIL_TOKEN_JSON")
    if not token_json:
        print("⚠️  GMAIL_TOKEN_JSON no configurado — correo no enviado")
        return

    token_data = json.loads(token_json)
    creds = google.oauth2.credentials.Credentials(
        token=token_data.get("token"),
        refresh_token=token_data.get("refresh_token"),
        token_uri=token_data.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=token_data.get("client_id"),
        client_secret=token_data.get("client_secret"),
        scopes=token_data.get("scopes", ["https://www.googleapis.com/auth/gmail.send"]),
    )

    service = build("gmail", "v1", credentials=creds)

    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"Seguimiento Legislativo — {fecha}"
    msg["To"] = ", ".join(destinatarios)
    msg["From"] = "Francisco Arellano <farellano@observatorio.cl>"

    # Cuerpo HTML
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # Excel adjunto
    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{excel_path.name}"')
    msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    print(f"✉️  Correo enviado a: {', '.join(destinatarios)}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    solo_reporte = "--solo-reporte" in sys.argv
    fecha_hoy = datetime.date.today().strftime("%d/%m/%Y")

    print(f"\n{'='*60}")
    print(f"Seguimiento Legislativo — {fecha_hoy}")
    print(f"{'='*60}\n")

    # 1. Cargar estado anterior
    estado_anterior = {}
    if ESTADO_PATH.exists():
        with open(ESTADO_PATH) as f:
            estado_anterior = json.load(f)
        print(f"Estado anterior cargado ({len(estado_anterior)} boletines)\n")
    else:
        print("Primera ejecución — no hay estado anterior\n")

    # 2. Scraping y actualización del Excel
    if not solo_reporte:
        print("Actualizando Excel...")
        actualizar_excel(EXCEL_PATH)
        print()

    # 3. Leer estado nuevo del Excel ya actualizado
    estado_nuevo = leer_estado_excel(EXCEL_PATH)

    # 4. Guardar nuevo estado como referencia para la próxima semana
    with open(ESTADO_PATH, "w") as f:
        json.dump(estado_nuevo, f, ensure_ascii=False, indent=2)
    print(f"Estado guardado en {ESTADO_PATH}\n")

    # 5. Detectar cambios
    cambios = detectar_cambios(estado_anterior, estado_nuevo)
    n_cambios = sum(1 for c in cambios if c["hubo_cambio"])
    print(f"Cambios detectados: {n_cambios} de {len(cambios)} boletines\n")

    # 6. Construir y enviar correo
    destinatarios_raw = os.environ.get("DESTINATARIOS", "")
    destinatarios = [d.strip() for d in destinatarios_raw.split(",") if d.strip()]

    if not destinatarios:
        print("⚠️  Variable DESTINATARIOS no configurada — correo no enviado")
        print("   Configura: export DESTINATARIOS='tu@email.cl,otro@email.cl'")
    else:
        html = construir_html(cambios, fecha_hoy)
        enviar_correo(html, EXCEL_PATH, destinatarios, fecha_hoy)

    print("\n✓ Proceso completado")


if __name__ == "__main__":
    main()
